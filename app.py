from flask import Flask, render_template, request, session, redirect, url_for, send_file
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.utils import secure_filename
import sqlite3
import os
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import requests
from bs4 import BeautifulSoup

app = Flask(__name__)
app.secret_key = 'your-secret-key-misdinar-2024'

DB_NAME = 'misdinar.db'
UPLOAD_DOKUMEN = 'static/uploads/dokumen'
UPLOAD_GALERI = 'static/uploads/galeri'

os.makedirs(UPLOAD_DOKUMEN, exist_ok=True)
os.makedirs(UPLOAD_GALERI, exist_ok=True)

def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    conn.execute('''CREATE TABLE IF NOT EXISTS users (username TEXT PRIMARY KEY, password TEXT NOT NULL, nama TEXT NOT NULL, role TEXT NOT NULL)''')
    
    db_columns = [col[1] for col in conn.execute('PRAGMA table_info(users)').fetchall()]
    if 'no_hp' not in db_columns:
        conn.execute('ALTER TABLE users ADD COLUMN no_hp TEXT')
    if 'nama_panggilan' not in db_columns:
        conn.execute('ALTER TABLE users ADD COLUMN nama_panggilan TEXT')
    
    conn.execute('''CREATE TABLE IF NOT EXISTS pengumuman (id INTEGER PRIMARY KEY AUTOINCREMENT, judul TEXT NOT NULL, waktu_pelaksanaan TEXT NOT NULL, tanggal_dibuat TEXT NOT NULL, target TEXT NOT NULL, pembuat TEXT NOT NULL)''')
    
    db_columns_peng = [col[1] for col in conn.execute('PRAGMA table_info(pengumuman)').fetchall()]
    if 'deskripsi' not in db_columns_peng:
        conn.execute('ALTER TABLE pengumuman ADD COLUMN deskripsi TEXT DEFAULT ""')

    conn.execute('''CREATE TABLE IF NOT EXISTS jadwal (id INTEGER PRIMARY KEY AUTOINCREMENT, jadwal_datetime TIMESTAMP NOT NULL, tanggal TEXT NOT NULL, bulan TEXT NOT NULL, hari TEXT NOT NULL, waktu TEXT NOT NULL, acara TEXT NOT NULL, status TEXT NOT NULL, pengguna TEXT NOT NULL, nama_pengguna TEXT NOT NULL, FOREIGN KEY(pengguna) REFERENCES users(username))''')
    conn.execute('''CREATE TABLE IF NOT EXISTS dokumen (id INTEGER PRIMARY KEY AUTOINCREMENT, judul TEXT NOT NULL, file_path TEXT NOT NULL, tanggal_dibuat TEXT NOT NULL, target TEXT NOT NULL, pembuat TEXT NOT NULL)''')
    conn.execute('''CREATE TABLE IF NOT EXISTS galeri (id INTEGER PRIMARY KEY AUTOINCREMENT, judul TEXT NOT NULL, foto_path TEXT NOT NULL, tanggal_dibuat TEXT NOT NULL, target TEXT NOT NULL, pembuat TEXT NOT NULL)''')
    
    users = [
        ('superadmin', 'super123', 'Super Admin', 'super admin', '081234567890'),
        ('abi', 'abi123', 'Abi', 'bph', '081111111111'),
        ('florensia', 'flor123', 'Florensia', 'bph', '082222222222'),
        ('angela_d', 'angela123', 'Angela D', 'bph', '083333333333'),
        ('rena', 'rena123', 'Rena', 'pelatihan', '084444444444'),
        ('gabriel', 'gabriel123', 'Gabriel', 'pelatihan', '085555555555'),
        ('dicky', 'dicky123', 'Dicky', 'pelatihan', '086666666666'),
        ('marlon@gmail.com', 'marlon123', 'Marlon', 'penjadwalan', '087777777777'),
        ('lydia', 'lydia123', 'Lydia', 'penjadwalan', '088888888888'),
        ('emily', 'emily123', 'Emily', 'penjadwalan', '089999999999'),
        ('galant', 'galant123', 'Galant', 'user', '081010101010'),
        ('dyota', 'dyota123', 'Dyota', 'user', '081212121212'),
        ('anton', '123456', 'Anton Wijaya', 'user', '081313131313'),
        ('budi', 'password', 'Budi Santoso', 'user', '081414141414'),
        ('citra', 'citra123', 'Citra Dewi', 'user', '081515151515')
    ]
    user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    if user_count == 0:
        for u in users:
            conn.execute('INSERT INTO users (username, password, nama, role, no_hp, nama_panggilan) VALUES (?, ?, ?, ?, ?, ?)', (u[0], u[1], u[2], u[3], u[4], u[2].split()[0]))
    
    conn.commit()
    conn.close()

init_db()

def get_quote_hari_ini():
    quotes = [
        "Melayani dengan hati, bukan karena ingin dipuji.",
        "Kasih itu sabar; kasih itu murah hati.",
        "Lakukan segala pekerjaanmu dalam kasih.",
        "Iman tanpa perbuatan adalah mati.",
        "Janganlah hendaknya kamu kuatir tentang apapun juga.",
        "Segala perkara dapat kutanggung di dalam Dia.",
        "Bersukacitalah senantiasa, tetaplah berdoa.",
        "Jadilah terang di tempat yang gelap."
    ]
    day_of_year = datetime.now().timetuple().tm_yday
    return quotes[day_of_year % len(quotes)]

def get_filtered_items(table_name, user_id=None):
    conn = get_db_connection()
    rows = conn.execute(f'SELECT * FROM {table_name} ORDER BY id DESC').fetchall()
    conn.close()
    
    filtered = []
    for r in rows:
        p = dict(r)
        target_list = p.get('target', 'semua').split(',')
        pembuat = p.get('pembuat', '')
        if 'semua' in target_list: filtered.append(p)
        elif user_id and (user_id in target_list or user_id == pembuat): filtered.append(p)
    return filtered

def get_jadwal_from_db():
    conn = get_db_connection()
    threshold_time = (datetime.now() - timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
    rows = conn.execute('SELECT * FROM jadwal WHERE jadwal_datetime >= ? ORDER BY jadwal_datetime ASC', (threshold_time,)).fetchall()
    conn.close()
    return [dict(row) for row in rows]

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        conn = get_db_connection()
        user = conn.execute('SELECT * FROM users WHERE username = ? AND password = ?', (username, password)).fetchone()
        conn.close()
        
        if user:
            session['user_id'] = user['username']
            session['nama_user'] = user['nama']
            session['role'] = user['role']
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Username atau password salah')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/')
def index():
    user_id = session.get('user_id')
    filtered_pengumuman = get_filtered_items('pengumuman', user_id)
    valid_jadwal = get_jadwal_from_db()
    quote_hari_ini = get_quote_hari_ini()
    
    if user_id:
        user_jadwal = [j for j in valid_jadwal if j['pengguna'] == user_id and j['status'] == 'Bertugas']
        return render_template('index.html', pengumuman=filtered_pengumuman, jadwal=user_jadwal[:4], 
                               quote=quote_hari_ini, user_id=user_id, is_logged_in=True)
    else:
        return render_template('index.html', pengumuman=filtered_pengumuman, jadwal=valid_jadwal[:8], 
                               quote=quote_hari_ini, is_logged_in=False)

@app.route('/jadwal')
def jadwal():
    valid_jadwal = get_jadwal_from_db()
    if 'user_id' in session:
        user_jadwal = [j for j in valid_jadwal if j['pengguna'] == session['user_id']]
        return render_template('jadwal.html', jadwal=user_jadwal, user_id=session['user_id'], is_logged_in=True, view_mode='private')
    else:
        return render_template('jadwal.html', jadwal=valid_jadwal, is_logged_in=False, view_mode='public')

@app.route('/anggota')
@login_required
def anggota():
    if session.get('role') not in ['super admin', 'bph', 'penjadwalan', 'pelatihan']: return redirect(url_for('index'))
    conn = get_db_connection()
    all_users = conn.execute("SELECT * FROM users WHERE role != 'super admin' ORDER BY nama ASC").fetchall()
    conn.close()
    return render_template('anggota.html', users=all_users, current_role=session.get('role'))

@app.route('/ubah-role', methods=['POST'])
@login_required
def ubah_role():
    current_role = session.get('role')
    if current_role not in ['super admin', 'bph']: return redirect(url_for('index'))
    
    username = request.form.get('username')
    new_role = request.form.get('role')
    new_password = request.form.get('password') 
    
    valid_roles = ['user', 'bph', 'penjadwalan', 'pelatihan']
    if current_role == 'super admin': valid_roles.append('super admin')
    
    if username and new_role in valid_roles:
        conn = get_db_connection()
        target_user = conn.execute("SELECT role FROM users WHERE username = ?", (username,)).fetchone()
        
        if target_user and target_user['role'] == 'super admin' and current_role != 'super admin':
            conn.close()
            return redirect(url_for('anggota'))
            
        if current_role == 'super admin' and new_password:
            conn.execute('UPDATE users SET role = ?, password = ? WHERE username = ?', (new_role, new_password, username))
        else:
            conn.execute('UPDATE users SET role = ? WHERE username = ?', (new_role, username))
            
        conn.commit()
        conn.close()
    return redirect(url_for('anggota'))

@app.route('/bph')
@login_required
def bph():
    if session.get('role') not in ['super admin', 'bph']: return redirect(url_for('index'))
    return render_template('bph.html')

@app.route('/penjadwalan', methods=['GET', 'POST'])
@login_required
def penjadwalan():
    if session.get('role') not in ['super admin', 'penjadwalan']: return redirect(url_for('index'))
    conn = get_db_connection()
    users_db = conn.execute("SELECT username, nama FROM users WHERE role != 'super admin' ORDER BY nama ASC").fetchall()
    conn.close()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # 1. RUTE BARU: Menampilkan Web Spreadsheet (Live Editor)
        if action == 'editor_web':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            if not start_date_str or not end_date_str: return redirect(url_for('penjadwalan'))
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            days_id = {0: 'Senin', 1: 'Selasa', 2: 'Rabu', 3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'}
            
            editor_data = []
            current_date = start_date
            while current_date <= end_date:
                hari_idx = current_date.weekday()
                hari = days_id[hari_idx]
                tgl_str = current_date.strftime('%Y-%m-%d')
                
                masses = []
                if hari_idx == 5: masses = [('05.30', 'Misa Pagi'), ('17.00', 'Misa Vigili')]
                elif hari_idx == 6: masses = [('06.00', 'Misa Pagi'), ('08.00', 'Misa Pagi II'), ('10.00', 'Misa Siang'), ('17.00', 'Misa Sore')]
                else: masses = [('05.30', 'Misa Pagi'), ('18.00', 'Misa Sore')]
                
                for waktu, acara in masses:
                    editor_data.append({
                        'tanggal': tgl_str,
                        'hari': hari,
                        'waktu': waktu,
                        'acara': acara
                    })
                current_date += timedelta(days=1)
            
            return render_template('penjadwalan.html', users=users_db, editor_data=editor_data, start_date=start_date_str, end_date=end_date_str)
            
        # 2. RUTE BARU: Menyimpan Hasil Ketikan dari Web Spreadsheet
        elif action == 'simpan_editor_web':
            tanggals = request.form.getlist('tgl[]')
            haris = request.form.getlist('hri[]')
            waktus = request.form.getlist('wkt[]')
            acaras = request.form.getlist('acr[]')
            petugas_list = request.form.getlist('ptgs[]')
            
            conn = get_db_connection()
            months_id = {1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MEI', 6: 'JUNI', 7: 'JULI', 8: 'AGUST', 9: 'SEPT', 10: 'OKT', 11: 'NOV', 12: 'DES'}
            
            for i in range(len(tanggals)):
                ptgs_raw = petugas_list[i]
                if not ptgs_raw or not ptgs_raw.strip(): continue # Lewati baris yang tidak diisi petugas
                
                tgl_raw = tanggals[i]
                waktu = waktus[i]
                acara = acaras[i]
                hari = haris[i]
                
                dt = datetime.strptime(tgl_raw, '%Y-%m-%d')
                jadwal_dt = datetime.strptime(f"{tgl_raw} {waktu.replace('.', ':')}:00", '%Y-%m-%d %H:%M:%S')
                
                # Memproses jika diketik banyak nama pakai koma
                list_petugas = [p.strip() for p in ptgs_raw.split(',')]
                for p in list_petugas:
                    if not p: continue
                    user_db = conn.execute("SELECT nama FROM users WHERE username=?", (p,)).fetchone()
                    nama_pengguna = user_db['nama'] if user_db else p
                    conn.execute('INSERT INTO jadwal (jadwal_datetime, tanggal, bulan, hari, waktu, acara, status, pengguna, nama_pengguna) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', 
                                 (jadwal_dt.strftime('%Y-%m-%d %H:%M:%S'), dt.strftime('%d'), months_id[dt.month], hari, waktu, acara, 'Bertugas', p, nama_pengguna))
            
            conn.commit()
            conn.close()
            return redirect(url_for('cetak_jadwal'))

        # 3. Rute Lama: Download Excel
        elif action == 'generate':
            start_date_str = request.form.get('start_date')
            end_date_str = request.form.get('end_date')
            if not start_date_str or not end_date_str: return redirect(url_for('penjadwalan'))
            
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            days_id = {0: 'Senin', 1: 'Selasa', 2: 'Rabu', 3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'}
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Form_Jadwal_Misdinar"
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2D5F3F")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            headers = ['Tanggal', 'Hari', 'Pukul', 'Acara Misa', 'Pilih Username (Gunakan Dropdown)']
            ws.append(headers)
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            row_idx = 2
            current_date = start_date
            while current_date <= end_date:
                hari_idx = current_date.weekday()
                hari = days_id[hari_idx]
                tgl_str = current_date.strftime('%Y-%m-%d')
                
                masses = []
                if hari_idx == 5: masses = [('05.30', 'Misa Pagi', 2), ('17.00', 'Misa Vigili', 9)]
                elif hari_idx == 6: masses = [('06.00', 'Misa Pagi', 9), ('08.00', 'Misa Pagi II', 9), ('10.00', 'Misa Siang', 9), ('17.00', 'Misa Sore', 9)]
                else: masses = [('05.30', 'Misa Pagi', 2), ('18.00', 'Misa Sore', 2)]
                
                for waktu, acara, slots in masses:
                    for _ in range(slots):
                        ws.cell(row=row_idx, column=1, value=tgl_str).alignment = center_align
                        ws.cell(row=row_idx, column=2, value=hari).alignment = center_align
                        ws.cell(row=row_idx, column=3, value=waktu).alignment = center_align
                        ws.cell(row=row_idx, column=4, value=acara).alignment = center_align
                        ws.cell(row=row_idx, column=5, value="")
                        for col in range(1, 6): ws.cell(row=row_idx, column=col).border = thin_border
                        row_idx += 1
                current_date += timedelta(days=1)
            
            user_list = [u['username'] for u in users_db]
            ws_users = wb.create_sheet("Data_Users")
            ws_users.sheet_state = 'hidden'
            for i, u in enumerate(user_list, 1): ws_users.cell(row=i, column=1, value=u)
            
            dv = DataValidation(type="list", formula1=f"Data_Users!$A$1:$A${len(user_list)}", allow_blank=True)
            dv.error ='Silakan pilih username dari daftar (atau ketik persis sesuai list).'
            dv.errorTitle = 'Username Tidak Valid'
            ws.add_data_validation(dv)
            dv.add(f'E2:E{row_idx-1}')
            
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 40
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, download_name=f'Form_Jadwal_{start_date_str}.xlsx', as_attachment=True)
            
        # 4. Rute Lama: Upload Excel    
        elif action == 'upload':
            file = request.files.get('file_excel')
            if file and file.filename.endswith('.xlsx'):
                try:
                    df = pd.read_excel(file)
                    conn = get_db_connection()
                    months_id = {1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MEI', 6: 'JUNI', 7: 'JULI', 8: 'AGUST', 9: 'SEPT', 10: 'OKT', 11: 'NOV', 12: 'DES'}
                    for index, row in df.iterrows():
                        col_petugas = [c for c in df.columns if 'user' in c.lower() or 'petugas' in c.lower()][0]
                        petugas_raw = str(row.get(col_petugas, ''))
                        if pd.isna(row.get(col_petugas)) or petugas_raw.strip() == '' or petugas_raw.lower() == 'nan': continue
                        tgl_raw = str(row[df.columns[0]]).split(' ')[0] 
                        waktu = str(row[df.columns[2]]).replace(':', '.') 
                        acara = str(row[df.columns[3]])
                        hari = str(row[df.columns[1]])
                        dt = datetime.strptime(tgl_raw, '%Y-%m-%d')
                        jadwal_dt = datetime.strptime(f"{tgl_raw} {waktu.replace('.', ':')}:00", '%Y-%m-%d %H:%M:%S')
                        list_petugas = [p.strip() for p in petugas_raw.split(',')]
                        for p in list_petugas:
                            if not p: continue
                            user_db = conn.execute("SELECT nama FROM users WHERE username=?", (p,)).fetchone()
                            nama_pengguna = user_db['nama'] if user_db else p
                            conn.execute('INSERT INTO jadwal (jadwal_datetime, tanggal, bulan, hari, waktu, acara, status, pengguna, nama_pengguna) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)', 
                                         (jadwal_dt.strftime('%Y-%m-%d %H:%M:%S'), dt.strftime('%d'), months_id[dt.month], hari, waktu, acara, 'Bertugas', p, nama_pengguna))
                    conn.commit()
                    conn.close()
                    return redirect(url_for('cetak_jadwal'))
                except Exception as e:
                    print("Error parsing Excel:", e)
                    return redirect(url_for('penjadwalan'))
                    
    return render_template('penjadwalan.html', users=users_db)

@app.route('/cetak-jadwal')
@login_required
def cetak_jadwal():
    if session.get('role') not in ['super admin', 'penjadwalan', 'bph']: return redirect(url_for('index'))
    
    conn = get_db_connection()
    threshold_time = datetime.now().strftime('%Y-%m-%d 00:00:00')
    rows = conn.execute('SELECT * FROM jadwal WHERE jadwal_datetime >= ? ORDER BY jadwal_datetime ASC', (threshold_time,)).fetchall()
    conn.close()

    jadwal_dict = {}
    for r in rows:
        dt = datetime.strptime(r['jadwal_datetime'], '%Y-%m-%d %H:%M:%S')
        date_key = dt.strftime('%Y-%m-%d')
        if date_key not in jadwal_dict:
            jadwal_dict[date_key] = {
                'hari': r['hari'], 
                'tanggal_teks': f"{r['hari']}, {dt.day} {r['bulan'].capitalize()} {dt.year}", 
                'misa': {}
            }
        
        wkt = r['waktu']
        if wkt not in jadwal_dict[date_key]['misa']: 
            jadwal_dict[date_key]['misa'][wkt] = []
        jadwal_dict[date_key]['misa'][wkt].append(r['nama_pengguna'])

    weeks = []
    current_week = {'senin_kamis': [], 'jumat_minggu': []}
    
    sorted_dates = sorted(jadwal_dict.keys())
    for d in sorted_dates:
        dt = datetime.strptime(d, '%Y-%m-%d')
        day_idx = dt.weekday()
        max_slots = 9 if day_idx >= 4 else 4
        misa_data = jadwal_dict[d]['misa']
        formatted_misa = {}
        for wkt, ptgs in misa_data.items():
            padded = ptgs.copy()
            while len(padded) < max_slots: padded.append("")
            formatted_misa[wkt] = padded
            
        day_obj = {'tanggal_teks': jadwal_dict[d]['tanggal_teks'], 'misa': formatted_misa, 'max_slots': max_slots}
        
        if day_idx <= 3: 
            if day_idx == 0 and len(current_week['senin_kamis']) > 0: 
                weeks.append(current_week)
                current_week = {'senin_kamis': [], 'jumat_minggu': []}
            current_week['senin_kamis'].append(day_obj)
        else: 
            current_week['jumat_minggu'].append(day_obj)
            
    if current_week['senin_kamis'] or current_week['jumat_minggu']: 
        weeks.append(current_week)
        
    return render_template('cetak_jadwal.html', weeks=weeks)

@app.route('/pelatihan')
@login_required
def pelatihan():
    if session.get('role') not in ['super admin', 'pelatihan']: return redirect(url_for('index'))
    return render_template('pelatihan.html')

@app.route('/pengaturan', methods=['GET', 'POST'])
@login_required
def pengaturan():
    user_id = session.get('user_id')
    conn = get_db_connection()
    if request.method == 'POST':
        nama_lengkap = request.form.get('nama_lengkap')
        nama_panggilan = request.form.get('nama_panggilan')
        no_hp = request.form.get('no_hp')
        password_baru = request.form.get('password_baru')
        if password_baru and password_baru.strip() != '':
            conn.execute('UPDATE users SET nama=?, nama_panggilan=?, no_hp=?, password=? WHERE username=?', (nama_lengkap, nama_panggilan, no_hp, password_baru, user_id))
        else:
            conn.execute('UPDATE users SET nama=?, nama_panggilan=?, no_hp=? WHERE username=?', (nama_lengkap, nama_panggilan, no_hp, user_id))
        conn.commit()
        session['nama_user'] = nama_lengkap
        user_data = conn.execute('SELECT * FROM users WHERE username = ?', (user_id,)).fetchone()
        conn.close()
        return render_template('pengaturan.html', user=user_data, success="Profil berhasil diperbarui!")
    user_data = conn.execute('SELECT * FROM users WHERE username = ?', (user_id,)).fetchone()
    conn.close()
    return render_template('pengaturan.html', user=user_data)

@app.route('/pendaftaran')
def pendaftaran(): return render_template('pendaftaran.html')

def get_grouped_users():
    conn = get_db_connection()
    all_users = conn.execute("SELECT username, nama, role FROM users WHERE role != 'super admin'").fetchall()
    conn.close()
    users_by_role = {}
    for u in all_users:
        r = u['role']
        if r not in users_by_role: users_by_role[r] = []
        users_by_role[r].append(dict(u))
    return users_by_role

@app.route('/pengumuman', methods=['GET', 'POST'])
def pengumuman():
    user_id = session.get('user_id')
    user_role = session.get('role')
    if request.method == 'POST' and user_role in ['super admin', 'bph', 'penjadwalan', 'pelatihan']:
        action = request.form.get('action')
        if action in ['tambah', 'edit']:
            judul = request.form.get('judul')
            deskripsi = request.form.get('deskripsi', '')
            waktu_pelaksanaan = request.form.get('waktu_pelaksanaan')
            target_list = request.form.getlist('target')
            target_str = 'semua' if 'semua' in target_list else ','.join(target_list)
            months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            now = datetime.now()
            tanggal_dibuat = f"{now.day} {months[now.month - 1]} {now.year}"
            conn = get_db_connection()
            if action == 'tambah':
                conn.execute('INSERT INTO pengumuman (judul, deskripsi, waktu_pelaksanaan, tanggal_dibuat, target, pembuat) VALUES (?, ?, ?, ?, ?, ?)', (judul, deskripsi, waktu_pelaksanaan, tanggal_dibuat, target_str, user_id))
            else:
                p_id = request.form.get('id')
                conn.execute('UPDATE pengumuman SET judul=?, deskripsi=?, waktu_pelaksanaan=?, target=? WHERE id=?', (judul, deskripsi, waktu_pelaksanaan, target_str, p_id))
            conn.commit()
            conn.close()
        elif action == 'hapus':
            p_id = request.form.get('id')
            conn = get_db_connection()
            conn.execute('DELETE FROM pengumuman WHERE id=?', (p_id,))
            conn.commit()
            conn.close()
        return redirect(url_for('pengumuman'))
    return render_template('pengumuman.html', pengumuman=get_filtered_items('pengumuman', user_id), role=user_role, users_by_role=get_grouped_users())

@app.route('/dokumen', methods=['GET', 'POST'])
def dokumen():
    user_id = session.get('user_id')
    user_role = session.get('role')
    if request.method == 'POST' and user_role in ['super admin', 'bph', 'penjadwalan', 'pelatihan']:
        action = request.form.get('action')
        if action in ['tambah', 'edit']:
            judul = request.form.get('judul')
            target_list = request.form.getlist('target')
            target_str = 'semua' if 'semua' in target_list else ','.join(target_list)
            months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            now = datetime.now()
            tanggal_dibuat = f"{now.day} {months[now.month - 1]} {now.year}"
            files = request.files.getlist('file')
            file_paths = []
            for idx, file in enumerate(files):
                if file and file.filename != '':
                    filename = secure_filename(file.filename)
                    filename = f"{int(now.timestamp())}_{idx}_{filename}"
                    file.save(os.path.join(UPLOAD_DOKUMEN, filename))
                    file_paths.append(f"uploads/dokumen/{filename}")
            conn = get_db_connection()
            if action == 'tambah':
                file_paths_str = ','.join(file_paths)
                conn.execute('INSERT INTO dokumen (judul, file_path, tanggal_dibuat, target, pembuat) VALUES (?, ?, ?, ?, ?)', (judul, file_paths_str, tanggal_dibuat, target_str, user_id))
            else:
                p_id = request.form.get('id')
                if file_paths:
                    file_paths_str = ','.join(file_paths)
                    conn.execute('UPDATE dokumen SET judul=?, file_path=?, target=? WHERE id=?', (judul, file_paths_str, target_str, p_id))
                else:
                    conn.execute('UPDATE dokumen SET judul=?, target=? WHERE id=?', (judul, target_str, p_id))
            conn.commit()
            conn.close()
        elif action == 'hapus':
            p_id = request.form.get('id')
            conn = get_db_connection()
            conn.execute('DELETE FROM dokumen WHERE id=?', (p_id,))
            conn.commit()
            conn.close()
        return redirect(url_for('dokumen'))
    return render_template('dokumen.html', dokumen=get_filtered_items('dokumen', user_id), role=user_role, users_by_role=get_grouped_users())

@app.route('/galeri', methods=['GET', 'POST'])
def galeri():
    user_id = session.get('user_id')
    user_role = session.get('role')
    if request.method == 'POST' and user_role in ['super admin', 'bph', 'penjadwalan', 'pelatihan']:
        action = request.form.get('action')
        if action in ['tambah', 'edit']:
            judul = request.form.get('judul')
            target_list = request.form.getlist('target')
            target_str = 'semua' if 'semua' in target_list else ','.join(target_list)
            months = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
            now = datetime.now()
            tanggal_dibuat = f"{now.day} {months[now.month - 1]} {now.year}"
            fotos = request.files.getlist('foto')
            foto_paths = []
            for idx, foto in enumerate(fotos):
                if foto and foto.filename != '':
                    filename = secure_filename(foto.filename)
                    filename = f"{int(now.timestamp())}_{idx}_{filename}"
                    foto.save(os.path.join(UPLOAD_GALERI, filename))
                    foto_paths.append(f"uploads/galeri/{filename}")
            conn = get_db_connection()
            if action == 'tambah':
                foto_paths_str = ','.join(foto_paths)
                conn.execute('INSERT INTO galeri (judul, foto_path, tanggal_dibuat, target, pembuat) VALUES (?, ?, ?, ?, ?)', (judul, foto_paths_str, tanggal_dibuat, target_str, user_id))
            else:
                p_id = request.form.get('id')
                if foto_paths:
                    foto_paths_str = ','.join(foto_paths)
                    conn.execute('UPDATE galeri SET judul=?, foto_path=?, target=? WHERE id=?', (judul, foto_paths_str, target_str, p_id))
                else:
                    conn.execute('UPDATE galeri SET judul=?, target=? WHERE id=?', (judul, target_str, p_id))
            conn.commit()
            conn.close()
        elif action == 'hapus':
            p_id = request.form.get('id')
            conn = get_db_connection()
            conn.execute('DELETE FROM galeri WHERE id=?', (p_id,))
            conn.commit()
            conn.close()
        return redirect(url_for('galeri'))
    return render_template('galeri.html', galeri=get_filtered_items('galeri', user_id), role=user_role, users_by_role=get_grouped_users())

@app.route('/kontak')
def kontak():
    conn = get_db_connection()
    pengurus_db = conn.execute("SELECT nama, role, no_hp FROM users WHERE role IN ('bph', 'penjadwalan', 'pelatihan') ORDER BY role ASC, nama ASC").fetchall()
    conn.close()
    pengurus_by_role = {}
    for p in pengurus_db:
        r = p['role']
        if r not in pengurus_by_role: pengurus_by_role[r] = []
        pengurus_by_role[r].append(dict(p))
    return render_template('kontak.html', pengurus_by_role=pengurus_by_role)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)